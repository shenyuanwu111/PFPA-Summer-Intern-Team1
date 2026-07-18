import json
from pathlib import Path


OUT = Path("outputs/Asset_TiC_TTC_conversion.ipynb")


def lines(text):
    return [line + "\n" for line in text.strip().splitlines()]


def md(text):
    return {"cell_type": "markdown", "metadata": {}, "source": lines(text)}


def code(text):
    return {
        "cell_type": "code",
        "execution_count": None,
        "metadata": {},
        "outputs": [],
        "source": lines(text),
    }


nb = {
    "cells": [
        md(
            """
# Asset Sheet Generation Using KMV and TiC/TTC Formulas

This notebook creates the `Asset` sheet required by `TiC TTC conversion.xlsx` for the 10 assigned public companies.

It starts from the existing KMV/EM output files, then applies the formulas from the updated course PDF:

- Asset volatility: `sigma_A`
- Asset return: `eta_A = annualized_mean_log_asset_return + 0.5 * sigma_A^2`
- Drift adjustment: `R_A = abs(eta_A - 0.5 * sigma_A^2)`
- TiC Risk Score: `100 * sigma_A^2 / ln(A/D)^2`
- CCM: `sigma_A^2 / (R_A * ln(A/D))`
- mu: `ln(A/D) / R_A`
- PIT PD: closed-form formula from the PDF
- TTC PD and S&P rating: spreadsheet conversion tables
"""
        ),
        md(
            """
## Important Date Convention

`Last Date` is one common valuation date across all companies: the latest trading date shared by all daily output files.

`Last Statement Date` remains company-specific because each company files financial statements on different dates. For point-in-time modeling this is correct: every company uses the latest filing available on or before the common valuation date.
"""
        ),
        code(
            r"""
from pathlib import Path
from datetime import datetime
import math
import shutil

import numpy as np
import pandas as pd
import openpyxl

PROJECT_DIR = Path.cwd()
KMV_OUTPUT_DIR = PROJECT_DIR / "kmv_em_output"
CONVERSION_WORKBOOK = Path(r"C:\Users\wujie\Downloads\TiC TTC conversion.xlsx")

OUTPUT_DIR = KMV_OUTPUT_DIR
ASSET_CSV = OUTPUT_DIR / "asset_tic_ttc_output.csv"
ASSET_WORKBOOK = OUTPUT_DIR / "Asset_TiC_TTC_conversion_output.xlsx"

SIGMA_COL = "\u03c3"
ETA_COL = "\u03b7"
MU_COL = "\u00b5"

ASSET_COLUMNS = [
    "Company", "Symbol", "Shares Outstanding", "Last Date", "Last Price",
    "Last Statement Date", "Debt/Short Term", "Debt/Long Term", "Total Debt",
    "Interest Rate", SIGMA_COL, "A", "R", ETA_COL, "CCM", MU_COL,
    "TiC Risk Score", "DD", "EDF", "PIT PD", "TTC PD", "SP Rating", "Outlook"
]

TRADING_DAYS = 250


def norm_cdf(x):
    return 0.5 * math.erfc(-float(x) / math.sqrt(2.0))


def norm_sf(x):
    return 0.5 * math.erfc(float(x) / math.sqrt(2.0))


print("KMV output directory:", KMV_OUTPUT_DIR)
print("Conversion workbook:", CONVERSION_WORKBOOK)
"""
        ),
        md("## 1. Load KMV/EM Daily Outputs"),
        code(
            r"""
daily_files = sorted(KMV_OUTPUT_DIR.glob("*_kmv_daily.csv"))
if not daily_files:
    raise FileNotFoundError(f"No *_kmv_daily.csv files found in {KMV_OUTPUT_DIR}")

company_frames = {}
common_dates = None

for file in daily_files:
    df = pd.read_csv(file, parse_dates=["Date", "filedDate", "periodEnd"])
    if df.empty:
        continue
    ticker = df["Ticker"].dropna().iloc[0]
    df = df.sort_values("Date").reset_index(drop=True)
    company_frames[ticker] = df

    dates = set(df["Date"].dt.normalize())
    common_dates = dates if common_dates is None else common_dates.intersection(dates)

if not common_dates:
    raise ValueError("No common trading date exists across the company daily files.")

COMMON_VALUATION_DATE = max(common_dates)

latest_rows = []
for ticker, df in company_frames.items():
    row = df[df["Date"].dt.normalize() == COMMON_VALUATION_DATE].iloc[-1].copy()
    latest_rows.append(row)

latest = pd.DataFrame(latest_rows).sort_values("Ticker").reset_index(drop=True)

print(f"Loaded {len(latest)} companies")
print("Common valuation date:", COMMON_VALUATION_DATE.date())
display(latest[[
    "Ticker", "Company", "Date", "AdjClose", "SharesOutstanding",
    "CurrentDebt", "LongTermDebt", "DefaultPoint", "RiskFreeRate",
    "AssetValue", "SigmaAsset", "DistanceToDefault", "TheoreticalPD"
]])
"""
        ),
        md("## 2. Calculate Asset Return and TiC Variables"),
        code(
            r"""
def annualized_mean_log_asset_return(df, valuation_date):
    history = df[df["Date"].dt.normalize() <= valuation_date].copy()
    asset_log_returns = np.log(history["AssetValue"]).diff().replace([np.inf, -np.inf], np.nan).dropna()
    if asset_log_returns.empty:
        return np.nan
    return float(asset_log_returns.mean() * TRADING_DAYS)


mean_log_return_by_ticker = {
    ticker: annualized_mean_log_asset_return(df, COMMON_VALUATION_DATE)
    for ticker, df in company_frames.items()
}

asset = pd.DataFrame()
asset["Company"] = latest["Company"]
asset["Symbol"] = latest["Ticker"]
asset["Shares Outstanding"] = latest["SharesOutstanding"]
asset["Last Date"] = latest["Date"].dt.date
asset["Last Price"] = latest["AdjClose"]
asset["Last Statement Date"] = latest["filedDate"].dt.date
asset["Debt/Short Term"] = latest["CurrentDebt"]
asset["Debt/Long Term"] = latest["LongTermDebt"]

# Use the KMV default point as Total Debt so the Asset sheet displays the same D used in formulas.
# D = short-term debt + 0.5 * long-term debt.
asset["Total Debt"] = latest["DefaultPoint"]
asset["KMV Default Point D"] = latest["DefaultPoint"]
asset["Interest Rate"] = latest["RiskFreeRate"]
asset[SIGMA_COL] = latest["SigmaAsset"]
asset["A"] = latest["AssetValue"]
asset["Annualized Mean Log Asset Return"] = asset["Symbol"].map(mean_log_return_by_ticker)
asset[ETA_COL] = asset["Annualized Mean Log Asset Return"] + 0.5 * asset[SIGMA_COL] ** 2

log_asset_to_debt = np.log(asset["A"] / asset["KMV Default Point D"])
sigma = asset[SIGMA_COL]
eta = asset[ETA_COL]
signed_drift = eta - 0.5 * sigma**2

# PDF notation:
# R_A = abs(eta_A - 0.5 * sigma_A^2)
# CCM = sigma_A^2 / (R_A * ln(A / D))
# mu = ln(A / D) / R_A
asset["R"] = signed_drift.abs()
asset["CCM"] = sigma**2 / (asset["R"] * log_asset_to_debt)
asset[MU_COL] = log_asset_to_debt / asset["R"]
asset["TiC Risk Score"] = 100.0 * sigma**2 / (log_asset_to_debt**2)

# DD uses the signed drift term from the KMV formula.
asset["DD"] = (log_asset_to_debt + signed_drift) / sigma
asset["EDF"] = asset["DD"].apply(norm_sf)


def pit_pd(ccm, mu):
    if pd.isna(ccm) or pd.isna(mu) or ccm <= 0 or mu <= 0:
        return np.nan
    sqrt_ccm = math.sqrt(float(ccm))
    sqrt_mu = math.sqrt(float(mu))
    term_1 = norm_cdf((1.0 / sqrt_ccm) * (1.0 / sqrt_mu - sqrt_mu))
    term_2 = math.exp(2.0 / float(ccm)) * norm_cdf(-(1.0 / sqrt_ccm) * (1.0 / sqrt_mu + sqrt_mu))
    return term_1 + term_2


asset["PIT PD"] = [
    pit_pd(ccm, mu)
    for ccm, mu in zip(asset["CCM"], asset[MU_COL])
]

display(asset[[
    "Symbol", "A", "KMV Default Point D", SIGMA_COL, ETA_COL,
    "R", "CCM", MU_COL, "TiC Risk Score", "DD", "EDF", "PIT PD"
]])
"""
        ),
        md("## 3. Load TTC and S&P Conversion Tables"),
        code(
            r"""
def load_lookup_table(workbook_path, sheet_name):
    raw = pd.read_excel(workbook_path, sheet_name=sheet_name, header=None)
    mu_values = pd.to_numeric(raw.iloc[1, 1:], errors="coerce")
    ccm_values = pd.to_numeric(raw.iloc[2:, 0], errors="coerce")
    values = raw.iloc[2:, 1:].apply(pd.to_numeric, errors="coerce")

    valid_rows = ccm_values.notna()
    valid_cols = mu_values.notna()

    ccm = ccm_values[valid_rows].to_numpy(dtype=float)
    mu = mu_values[valid_cols].to_numpy(dtype=float)
    grid = values.loc[valid_rows, valid_cols.to_numpy()].to_numpy(dtype=float)
    return ccm, mu, grid


def bilinear_lookup(x, y, x_axis, y_axis, grid):
    x_original, y_original = float(x), float(y)
    x = float(np.clip(x_original, x_axis.min(), x_axis.max()))
    y = float(np.clip(y_original, y_axis.min(), y_axis.max()))

    xi = np.searchsorted(x_axis, x, side="right") - 1
    yi = np.searchsorted(y_axis, y, side="right") - 1
    xi = int(np.clip(xi, 0, len(x_axis) - 2))
    yi = int(np.clip(yi, 0, len(y_axis) - 2))

    x0, x1 = x_axis[xi], x_axis[xi + 1]
    y0, y1 = y_axis[yi], y_axis[yi + 1]
    q00 = grid[xi, yi]
    q01 = grid[xi, yi + 1]
    q10 = grid[xi + 1, yi]
    q11 = grid[xi + 1, yi + 1]

    wx = 0.0 if x1 == x0 else (x - x0) / (x1 - x0)
    wy = 0.0 if y1 == y0 else (y - y0) / (y1 - y0)
    value = (
        q00 * (1 - wx) * (1 - wy)
        + q10 * wx * (1 - wy)
        + q01 * (1 - wx) * wy
        + q11 * wx * wy
    )
    clipped = (x != x_original) or (y != y_original)
    return value, clipped


ttc_ccm, ttc_mu, ttc_grid = load_lookup_table(CONVERSION_WORKBOOK, "TTC")
print("TTC table CCM range:", ttc_ccm.min(), "to", ttc_ccm.max())
print("TTC table mu range:", ttc_mu.min(), "to", ttc_mu.max())

sp = pd.read_excel(CONVERSION_WORKBOOK, sheet_name="SP", header=None, usecols=[0, 1])
sp = sp.iloc[1:].copy()
sp.columns = ["SP Rating", "S&P PD"]
sp["S&P PD"] = pd.to_numeric(sp["S&P PD"], errors="coerce")
sp = sp.dropna(subset=["SP Rating", "S&P PD"]).sort_values("S&P PD").reset_index(drop=True)


def map_sp_rating(pd_value):
    value = float(pd_value)
    thresholds = sp["S&P PD"].to_numpy(dtype=float)
    ratings = sp["SP Rating"].astype(str).to_numpy()
    idx = np.searchsorted(thresholds, value, side="right") - 1
    idx = int(np.clip(idx, 0, len(ratings) - 1))
    return ratings[idx]


display(sp.head(12))
"""
        ),
        md("## 4. Calculate TTC PD and S&P Rating"),
        code(
            r"""
ttc_values = []
lookup_clipped = []
for ccm, mu in zip(asset["CCM"], asset[MU_COL]):
    value, clipped = bilinear_lookup(ccm, mu, ttc_ccm, ttc_mu, ttc_grid)
    ttc_values.append(value)
    lookup_clipped.append(clipped)

asset["TTC PD"] = ttc_values
asset["SP Rating"] = asset["TTC PD"].apply(map_sp_rating)
asset["Outlook"] = asset["TTC PD"] - asset["PIT PD"]
asset["TTC Lookup Clipped"] = lookup_clipped

display(asset[[
    "Symbol", "CCM", MU_COL, "PIT PD", "TTC PD",
    "SP Rating", "Outlook", "TTC Lookup Clipped"
]])
"""
        ),
        md("## 5. Save Asset Outputs"),
        code(
            r"""
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

asset_output = asset[ASSET_COLUMNS].copy()
asset_output.to_csv(ASSET_CSV, index=False)

try:
    shutil.copyfile(CONVERSION_WORKBOOK, ASSET_WORKBOOK)
    final_workbook = ASSET_WORKBOOK
except PermissionError:
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    final_workbook = ASSET_WORKBOOK.with_name(f"{ASSET_WORKBOOK.stem}_{timestamp}{ASSET_WORKBOOK.suffix}")
    shutil.copyfile(CONVERSION_WORKBOOK, final_workbook)
    print("Default workbook was locked, saved a timestamped copy instead.")

wb = openpyxl.load_workbook(final_workbook)
ws = wb["Asset"]

if ws.max_row > 1:
    ws.delete_rows(2, ws.max_row - 1)

for col_idx, col_name in enumerate(ASSET_COLUMNS, start=1):
    ws.cell(row=1, column=col_idx).value = col_name

for row_idx, row in enumerate(asset_output.itertuples(index=False), start=2):
    for col_idx, value in enumerate(row, start=1):
        ws.cell(row=row_idx, column=col_idx).value = None if pd.isna(value) else value

wb.save(final_workbook)

print("Saved CSV:", ASSET_CSV)
print("Saved workbook:", final_workbook)
"""
        ),
        md("## 6. Validation Summary"),
        code(
            r"""
required = set(ASSET_COLUMNS)
missing = [col for col in required if asset_output[col].isna().any()]

print("Rows:", len(asset_output))
print("Columns:", len(asset_output.columns))
print("Missing fields:", missing)
print("Common valuation date:", COMMON_VALUATION_DATE.date())
print("TTC lookup clipped rows:", int(asset["TTC Lookup Clipped"].sum()))

display(asset_output[[
    "Symbol", "Last Date", "Last Statement Date", SIGMA_COL, ETA_COL,
    "R", "CCM", MU_COL, "TiC Risk Score", "DD", "EDF",
    "PIT PD", "TTC PD", "SP Rating", "Outlook"
]])
"""
        ),
    ],
    "metadata": {
        "kernelspec": {
            "display_name": "Python 3",
            "language": "python",
            "name": "python3",
        },
        "language_info": {"name": "python", "pygments_lexer": "ipython3"},
    },
    "nbformat": 4,
    "nbformat_minor": 5,
}

OUT.parent.mkdir(parents=True, exist_ok=True)
OUT.write_text(json.dumps(nb, indent=2, ensure_ascii=True), encoding="utf-8")
print(f"Wrote {OUT}")
